"""
從 2026 PACKING LIST.xlsx 的「重量明細」sheet 擷取產品資料，產出 products.json
執行一次即可：python scripts/extract_data.py
"""
import json
import re
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "..",
    "2026 PACKING LIST.xlsx",
)
OUTPUT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data",
    "products.json",
)


def extract_number(value):
    """從混合文字中擷取數字，例如 '5 (滿箱)' → 5, '2 (SA.HS)' → 2"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"(\d+\.?\d*)", str(value))
    if match:
        return float(match.group(1))
    return None


def extract_products():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["重量明細"]

    products = {}
    current_model = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Column C = model name (index 2)
        model_cell = row[2].value if len(row) > 2 else None
        # Column I = boxes/sets per carton (index 8)
        sets_cell = row[8].value if len(row) > 8 else None
        # Column J = old gross weight (index 9)
        old_weight_cell = row[9].value if len(row) > 9 else None
        # Column N = new gross weight (index 13)
        new_weight_cell = row[13].value if len(row) > 13 else None

        # Update current model if column C has a value
        if model_cell is not None:
            model_name = str(model_cell).strip()
            # Skip multiline entries or empty
            if "\n" in model_name or not model_name:
                current_model = None
                continue
            current_model = model_name

        if current_model is None:
            continue

        # Parse sets per carton
        sets_per_carton = extract_number(sets_cell)
        if sets_per_carton is None or sets_per_carton <= 0:
            continue

        # Parse weight: prefer new package weight, fallback to old
        weight = extract_number(new_weight_cell)
        if weight is None or weight <= 0:
            weight = extract_number(old_weight_cell)
        if weight is None or weight <= 0:
            continue

        sets_per_carton = int(sets_per_carton)

        # Add to products dict
        if current_model not in products:
            products[current_model] = []

        # Avoid duplicate entries
        existing = [
            opt
            for opt in products[current_model]
            if opt["sets_per_carton"] == sets_per_carton
            and abs(opt["weight_kg"] - weight) < 0.01
        ]
        if not existing:
            products[current_model].append(
                {
                    "sets_per_carton": sets_per_carton,
                    "weight_kg": round(weight, 2),
                }
            )

    # Sort options by sets_per_carton for each model
    for model in products:
        products[model].sort(key=lambda x: x["sets_per_carton"])

    return products


def main():
    print("正在從 Excel 擷取產品資料...")
    products = extract_products()
    print(f"擷取完成：{len(products)} 個產品型號")

    # Ensure output directory exists
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    print(f"已儲存至 {OUTPUT_PATH}")

    # Print summary
    for model, options in sorted(products.items()):
        opts_str = ", ".join(
            [f"{o['sets_per_carton']}sets/{o['weight_kg']}kg" for o in options]
        )
        print(f"  {model}: {opts_str}")


if __name__ == "__main__":
    main()
