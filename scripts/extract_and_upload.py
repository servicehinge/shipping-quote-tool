"""
從修正後的 Excel 重新提取產品資料並上傳到 Google Sheets「產品資料」工作表

策略：
- 修正後的欄位 C（產品型號含 -X 後綴）來自「2026 PACKING LIST (欄位C修正).xlsx」
- 計算好的公式值（L, M, N, O, P）來自原始「2026 PACKING LIST.xlsx」
- 對於 N 值異常（L=None 導致 N ≈ M）的列，用 base_inner_weight * I + M 重新計算

工作表格式：
  產品型號 | sets_per_carton | weight_kg
"""

import re
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import config
from services.google_sheets import get_or_create_worksheet

EXCEL_DIR = os.path.join(os.path.dirname(__file__), "..", "..")
ORIGINAL_FILE = os.path.join(EXCEL_DIR, "2026 PACKING LIST.xlsx")
CORRECTED_FILE = os.path.join(EXCEL_DIR, "2026 PACKING LIST (欄位C修正).xlsx")

SKIP_PATTERNS = [
    "檔環", "DISPLAY", "Demo", "木門", "格柵",
    "ADA HANDLE", "背板", "SA1", "SA.", "HS.",
]

HEADER = ["產品型號", "sets_per_carton", "weight_kg"]


def parse_sets_per_carton(raw):
    """Parse Column I value → int, handling text like '5 (滿箱)'"""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    m = re.match(r"(\d+)", str(raw))
    return int(m.group(1)) if m else None


def should_skip(model: str) -> bool:
    """Check if model should be excluded"""
    for pat in SKIP_PATTERNS:
        if pat in model:
            return True
    # Skip multiline or obviously non-product entries
    if "\n" in model or len(model) > 30:
        return True
    return False


def extract_products():
    """Extract product data from Excel files"""
    print("讀取 Excel 檔案...")
    wb_orig = openpyxl.load_workbook(ORIGINAL_FILE, data_only=True)
    ws_orig = wb_orig["重量明細"]

    wb_corr = openpyxl.load_workbook(CORRECTED_FILE, data_only=False)
    ws_corr = wb_corr["重量明細"]

    max_row = ws_orig.max_row

    # ── Pass 1: Collect base inner weight per (model, G) from I=1 rows ──
    base_weights = {}  # (corrected_model, G) -> inner_box_weight
    for row_idx in range(3, max_row + 1):
        model_corr = ws_corr[f"C{row_idx}"].value
        if not model_corr or not isinstance(model_corr, str):
            continue
        model_corr = model_corr.strip()
        if should_skip(model_corr):
            continue

        g = ws_orig[f"G{row_idx}"].value
        i_val = parse_sets_per_carton(ws_orig[f"I{row_idx}"].value)
        l_val = ws_orig[f"L{row_idx}"].value

        if g is None or i_val is None or i_val != 1:
            continue
        if l_val is None or not isinstance(l_val, (int, float)):
            continue

        try:
            g_int = int(float(g))
        except (ValueError, TypeError):
            continue

        key = (model_corr, g_int)
        base_weights[key] = float(l_val)

    print(f"收集到 {len(base_weights)} 組基礎內盒重量")

    # ── Pass 2: Extract all valid rows ──
    rows = []
    fixed_count = 0
    skipped_count = 0

    for row_idx in range(3, max_row + 1):
        model_corr = ws_corr[f"C{row_idx}"].value
        if not model_corr or not isinstance(model_corr, str):
            continue
        model_corr = model_corr.strip()
        if should_skip(model_corr):
            continue

        g = ws_orig[f"G{row_idx}"].value
        i_val = parse_sets_per_carton(ws_orig[f"I{row_idx}"].value)

        if g is None or i_val is None or i_val <= 0:
            continue

        try:
            g_int = int(float(g))
        except (ValueError, TypeError):
            continue

        l_val = ws_orig[f"L{row_idx}"].value
        m_val = ws_orig[f"M{row_idx}"].value
        n_val = ws_orig[f"N{row_idx}"].value

        # Determine total weight per carton
        weight = None

        # Check if N is valid (> 1 kg typically means it includes product weight)
        if n_val is not None and isinstance(n_val, (int, float)) and n_val > 1:
            weight = float(n_val)
        else:
            # N is bad → recalculate from base inner weight
            key = (model_corr, g_int)
            base_w = base_weights.get(key)
            m_float = float(m_val) if m_val and isinstance(m_val, (int, float)) else 0

            if base_w:
                weight = base_w * i_val + m_float
                fixed_count += 1
            else:
                # Cannot determine weight - skip
                skipped_count += 1
                continue

        weight = round(weight, 2)
        if weight <= 0:
            skipped_count += 1
            continue

        rows.append([model_corr, i_val, weight])

    print(f"提取完成: {len(rows)} 筆有效資料")
    print(f"修復 N 值: {fixed_count} 筆")
    print(f"跳過無效: {skipped_count} 筆")

    # ── Deduplicate (same model + same sets_per_carton → keep first) ──
    seen = set()
    unique_rows = []
    for row in rows:
        key = (row[0], row[1])
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)
        # If duplicate, skip silently (same model+packing combo)

    if len(unique_rows) < len(rows):
        print(f"去重: {len(rows)} → {unique_rows.__len__()} 筆（移除 {len(rows) - len(unique_rows)} 筆重複）")

    # Sort by model name, then sets_per_carton
    unique_rows.sort(key=lambda r: (r[0], r[1]))

    return unique_rows


def upload_to_sheets(rows):
    """Upload extracted data to Google Sheets"""
    print(f"\n上傳到 Google Sheets「{config.SHEET_NAME_PRODUCTS}」...")
    ws = get_or_create_worksheet(
        config.SHEET_NAME_PRODUCTS, rows=len(rows) + 10, cols=5
    )
    ws.clear()
    ws.update([HEADER] + rows, value_input_option="USER_ENTERED")
    print(f"上傳完成！共 {len(rows)} 列（不含標題）")

    # Count unique models
    models = set(r[0] for r in rows)
    print(f"共 {len(models)} 個不同型號")


def main():
    rows = extract_products()

    # Show sample data
    print("\n=== 前 20 筆資料 ===")
    print(f"{'產品型號':<25} {'sets/箱':>8} {'重量kg':>10}")
    print("-" * 45)
    for r in rows[:20]:
        print(f"{r[0]:<25} {r[1]:>8} {r[2]:>10.2f}")

    # Show unique models
    models = sorted(set(r[0] for r in rows))
    print(f"\n=== 所有型號 ({len(models)} 個) ===")
    for m in models:
        opts = [r for r in rows if r[0] == m]
        opts_str = ", ".join(f"{r[1]}sets/{r[2]}kg" for r in opts)
        print(f"  {m}: {opts_str}")

    upload_to_sheets(rows)


if __name__ == "__main__":
    main()
